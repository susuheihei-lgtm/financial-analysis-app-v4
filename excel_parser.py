"""
Excelファイルからstock_data.json形式のデータを抽出するパーサー
シートが不足していても利用可能なデータだけで分析を行う
.xls（旧形式）および日本語ラベルの縦型レイアウトにも対応
"""
import os
import openpyxl
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


# ---------- xlrd → openpyxl 互換アダプタ ----------

class _XlrdCellAdapter:
    """xlrdのセル値をopenpyxlのcell.valueインターフェースで返す"""
    def __init__(self, value):
        self.value = value


class _XlrdSheetAdapter:
    """xlrdのシートをopenpyxlのワークシート互換で返す"""
    def __init__(self, xlrd_sheet):
        self._sheet = xlrd_sheet
        self.max_row = xlrd_sheet.nrows
        self.max_column = xlrd_sheet.ncols
        self.title = xlrd_sheet.name

    def cell(self, row, column):
        try:
            v = self._sheet.cell_value(row - 1, column - 1)
            if v == '':
                v = None
            return _XlrdCellAdapter(v)
        except IndexError:
            return _XlrdCellAdapter(None)


class _XlrdWorkbookAdapter:
    """xlrdのワークブックをopenpyxl互換で返す"""
    def __init__(self, xlrd_wb):
        self._wb = xlrd_wb
        self.sheetnames = xlrd_wb.sheet_names()

    def __getitem__(self, name):
        return _XlrdSheetAdapter(self._wb.sheet_by_name(name))


def _load_workbook(filepath):
    """拡張子に応じてopenpyxlまたはxlrdでワークブックを開く"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xls':
        if not HAS_XLRD:
            raise ImportError(
                '.xlsファイルの読み込みにはxlrdが必要です。'
                'pip install xlrd でインストールしてください。'
            )
        wb = xlrd.open_workbook(filepath)
        return _XlrdWorkbookAdapter(wb)
    else:
        return openpyxl.load_workbook(filepath, data_only=True)


# ---------- 日本語縦型レイアウト検出・パーサー ----------

# セクションヘッダー
_JP_SECTION_HEADERS = {'業績', '財務', 'CF', 'キャッシュフロー', '配当', '株価指標'}

# 日本語ラベル → 内部キーのマッピング
_JP_LABEL_MAP = {
    # 業績セクション
    '売上高': 'revenue',
    '営業利益': 'op_income',
    '経常利益': 'ordinary_income',
    '純利益': 'net_income',
    '当期純利益': 'net_income',
    'EPS': 'eps',
    '一株益': 'eps',
    '1株益': 'eps',
    'ROE': 'roe',
    'ROA': 'roa',
    '営業利益率': 'op_margin_pct',
    '営業CFマージン': 'ocf_margin_pct',
    # 財務セクション
    '総資産': 'total_assets',
    '純資産': 'net_assets',
    '株主資本': 'total_equity',
    '自己資本': 'total_equity',
    '利益剰余金': 'retained_earnings',
    '短期借入金': 'short_term_debt',
    '長期借入金': 'long_term_debt',
    '有利子負債': 'total_debt',
    'BPS': 'bps',
    '一株純資産': 'bps',
    '1株純資産': 'bps',
    '自己資本比率': 'equity_ratio_pct',
    # CFセクション
    '営業CF': 'ocf',
    '投資CF': 'investing_cf',
    '財務CF': 'financing_cf',
    '設備投資': 'capex',
    '現金同等物': 'cash',
    '現金及び現金同等物': 'cash',
    'フリーCF': 'fcf',
    # 配当セクション
    '一株配当': 'dividend_per_share',
    '1株配当': 'dividend_per_share',
    '配当性向': 'payout_ratio_pct',
    '総還元性向': 'total_return_ratio',
    '剰余金の配当': 'dividend_total',
    '自社株買い': 'buyback',
    '純資産配当率': 'doe',
}


def _is_japanese_vertical_layout(wb):
    """日本語の縦型レイアウト（1シートに業績・財務・CF等が並ぶ）かどうかを判定"""
    if len(wb.sheetnames) > 4:
        return False
    ws = wb[wb.sheetnames[0]]
    found_sections = set()
    for r in range(1, min(ws.max_row + 1, 50)):
        val = ws.cell(row=r, column=1).value
        if val is not None and str(val).strip() in _JP_SECTION_HEADERS:
            found_sections.add(str(val).strip())
    return len(found_sections) >= 2


def _parse_numeric(v):
    """値を数値に変換。'-'や空文字、'（予想）'等はNoneにする"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(',', '')
    if s in ('-', '－', '', '―', 'N/A', 'n/a'):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_japanese_vertical(wb):
    """日本語縦型レイアウトのExcelをパースしてanalyzer用のdictとts_dataを返す"""
    ws = wb[wb.sheetnames[0]]

    # 全行を読み込み
    rows = []
    for r in range(1, ws.max_row + 1):
        row = []
        for c in range(1, ws.max_column + 1):
            row.append(ws.cell(row=r, column=c).value)
        rows.append(row)

    # 会社名の検出（1行目）
    company_name = str(rows[0][0]).strip() if rows and rows[0][0] else ''

    # セクションごとにデータを読み取る
    # 各セクションはヘッダー行の次にラベル行（年度, 売上高, ...）、その後にデータ行が続く
    raw_data = {}  # key -> [値のリスト（古い→新しい順）]
    dates_by_section = {}
    current_section = None
    header_row = None
    skip_forecast = False

    for i, row in enumerate(rows):
        first_cell = str(row[0]).strip() if row[0] is not None else ''

        # セクションヘッダーを検出
        if first_cell in _JP_SECTION_HEADERS:
            current_section = first_cell
            header_row = None
            continue

        # セクション内のラベル行を検出（年度, 売上高, ...のような行）
        if current_section and header_row is None:
            if first_cell in ('年度', '決算期', '決算年度'):
                header_row = row
                continue
            continue

        # データ行を処理
        if current_section and header_row is not None:
            # 空行でセクション終了
            if not first_cell:
                current_section = None
                header_row = None
                continue

            # 予想行はスキップ
            last_cell = row[-1] if row else None
            if last_cell and '予想' in str(last_cell):
                continue

            # 年度を取得
            date_val = first_cell
            if current_section not in dates_by_section:
                dates_by_section[current_section] = []
            dates_by_section[current_section].append(date_val)

            # 各列のデータを対応するキーに格納
            for col_idx in range(1, len(header_row)):
                label = str(header_row[col_idx]).strip() if header_row[col_idx] else ''
                if not label or label == '':
                    continue
                key = _JP_LABEL_MAP.get(label)
                if key is None:
                    continue
                val = _parse_numeric(row[col_idx] if col_idx < len(row) else None)
                if key not in raw_data:
                    raw_data[key] = []
                raw_data[key].append(val)

    # 日付を取得（業績セクションを優先）
    dates_raw = dates_by_section.get('業績', dates_by_section.get(
        list(dates_by_section.keys())[0] if dates_by_section else '', []))

    # データは古い→新しい順で並んでいるので、新しい順に反転
    for key in raw_data:
        raw_data[key] = list(reversed(raw_data[key]))
    dates_raw = list(reversed(dates_raw))

    # ヘルパー
    def g(key, idx):
        lst = raw_data.get(key, [])
        return lst[idx] if idx < len(lst) else None

    def g_list(key):
        return raw_data.get(key, [])

    # ROE/ROAは%値として直接入っている場合がある（10.86 = 10.86%）
    # 既に%単位なのでto_pctは不要
    roe_list = g_list('roe')
    roa_list = g_list('roa')
    roe_now = g('roe', 0)
    roe_3y = g('roe', 2)
    roe_5y = g('roe', 4)
    roa_now = g('roa', 0)
    roa_3y = g('roa', 2)
    roa_5y = g('roa', 4)
    roe_growth = roe_now - roe_5y if (roe_now is not None and roe_5y is not None) else None

    # 自己資本比率（%値で直接入っている）
    equity_ratio = g('equity_ratio_pct', 0)
    equity_ratio_5y = g('equity_ratio_pct', 4)
    # equity_ratio_pctがなければ、株主資本/総資産から計算
    if equity_ratio is None and g('total_equity', 0) and g('total_assets', 0):
        equity_ratio = (g('total_equity', 0) / g('total_assets', 0)) * 100
    if equity_ratio_5y is None and g('total_equity', 4) and g('total_assets', 4):
        equity_ratio_5y = (g('total_equity', 4) / g('total_assets', 4)) * 100

    # 営業利益率の計算
    revenue = g_list('revenue')
    op_income = g_list('op_income')
    net_income = g_list('net_income')
    op_margin_vals = []
    if g_list('op_margin_pct'):
        op_margin_vals = g_list('op_margin_pct')[:5]
    elif op_income and revenue:
        for i in range(min(5, len(op_income))):
            oi = g('op_income', i)
            rev = g('revenue', i)
            if oi is not None and rev and rev != 0:
                op_margin_vals.append(round(oi / rev * 100, 2))
            else:
                op_margin_vals.append(None)

    # FCF計算（営業CF - 設備投資の絶対値）
    ocf_list = g_list('ocf')
    capex_list = g_list('capex')
    fcf_list = g_list('fcf')
    if not fcf_list and ocf_list and capex_list:
        fcf_list = []
        for i in range(min(len(ocf_list), len(capex_list))):
            o = ocf_list[i]
            c = capex_list[i]
            if o is not None and c is not None:
                fcf_list.append(o + c)  # capexは通常マイナス値
            else:
                fcf_list.append(None)

    # 配当性向（%値）
    payout_pct = g('payout_ratio_pct', 0)
    payout_pct_5y = g('payout_ratio_pct', 4)

    # NOPAT
    nopat = g('op_income', 0) * 0.75 if g('op_income', 0) else None
    nopat_5y = g('op_income', 4) * 0.75 if g('op_income', 4) else None

    data = {
        "company": company_name,
        "ticker": "",
        "industry": "製造・サービス",

        "revenue": [g('revenue', i) for i in range(min(5, len(revenue)))],
        "fcf": [fcf_list[i] if i < len(fcf_list) else None for i in range(min(5, len(fcf_list)))],
        "eps": [g('eps', i) for i in range(min(5, len(g_list('eps'))))],

        "roe": [roe_now, roe_3y, roe_5y],
        "roe_growth_rate": roe_growth,
        "roa": [roa_now, roa_3y, roa_5y],

        "equity_ratio": equity_ratio,
        "equity_ratio_5y": equity_ratio_5y,
        "quick_ratio": None,
        "quick_ratio_5y": None,
        "current_ratio": None,
        "current_ratio_5y": None,

        "operating_cf": [g('ocf', i) for i in range(min(5, len(ocf_list)))],
        "investing_cf": [g('investing_cf', i) for i in range(min(5, len(g_list('investing_cf'))))],
        "financing_cf": [g('financing_cf', i) for i in range(min(5, len(g_list('financing_cf'))))],
        "op_margin": op_margin_vals,
        "ebitda_margin": None,
        "ebitda_margin_5y": None,

        "debt_fcf": None,
        "debt_fcf_5y": None,
        "nd_ebitda": None,
        "ev": None,
        "per": None,
        "per_5y": None,
        "pbr": None,
        "pbr_5y": None,

        "nopat": nopat,
        "nopat_5y": nopat_5y,
        "invested_capital": None,
        "invested_capital_5y": None,
        "wacc": None,

        "accounts_receivable": None,
        "accounts_receivable_5y": None,
        "inventory": None,
        "inventory_5y": None,
        "accounts_payable": None,
        "accounts_payable_5y": None,
        "cogs": None,
        "cogs_5y": None,
        "sga_ratio": None,
        "sga_ratio_5y": None,

        "total_assets": g('total_assets', 0),
        "total_assets_5y": g('total_assets', 4),
        "fixed_assets": None,
        "fixed_assets_5y": None,
        "tangible_fixed_assets": None,
        "tangible_fixed_assets_5y": None,
        "intangible_fixed_assets": None,
        "intangible_fixed_assets_5y": None,

        "net_income_val": g('net_income', 0),
        "net_income_val_5y": g('net_income', 4),
        "op_income_val": g('op_income', 0),
        "op_income_val_5y": g('op_income', 4),
        "interest_exp": None,
        "interest_exp_5y": None,
        "other_exp": None,
        "other_exp_5y": None,
        "pretax_income": None,
        "pretax_income_5y": None,
        "income_tax": None,
        "income_tax_5y": None,
        "total_equity": g('total_equity', 0),
        "total_equity_5y": g('total_equity', 4),

        "dividend_yield": None,
        "dividend_yield_5y": None,
        "payout_ratio": payout_pct,
        "payout_ratio_5y": payout_pct_5y,

        "d1_mgmt_change": "○",
        "d2_ownership": "○",
        "d3_esg": "○",
    }

    # 時系列データ（チャート用）
    def to_pct(v):
        return v * 100 if v is not None else None

    date_strs = [str(d)[:4] if d else "" for d in dates_raw]
    investing_cf_list = g_list('investing_cf')
    financing_cf_list = g_list('financing_cf')
    eps_list = g_list('eps')

    ts_data = {
        "dates": date_strs,
        "revenue": list(revenue),
        "net_income": list(net_income),
        "fcf": list(fcf_list),
        "eps": list(eps_list),
        "ocf": list(ocf_list),
        "investing_cf": list(investing_cf_list),
        "financing_cf": list(financing_cf_list),
        "ebitda": [],
        "total_assets": list(g_list('total_assets')),
        "total_equity": list(g_list('total_equity')),
        "total_debt": list(g_list('total_debt')),
        "roe": list(roe_list),
        "roa": list(roa_list),
        "op_margin": list(op_margin_vals) + [None] * max(0, len(revenue) - len(op_margin_vals)),
        "quick_ratio": [],
        "current_ratio": [],
        "equity_ratio": list(g_list('equity_ratio_pct')),
        "ebitda_margin": [],
        "debt_fcf": [],
        "roic": [],
        "capex": list(capex_list),
        "sga": [],
        "da": [],
        "pe_ratio": [],
        "pb_ratio": [],
        "debt_ebitda": [],
        "nd_ebitda": [],
        "dividend_yield": [],
        "payout_ratio": list(g_list('payout_ratio_pct')),
    }

    # 営業利益率の時系列計算
    if not ts_data["op_margin"] or all(v is None for v in ts_data["op_margin"]):
        ts_data["op_margin"] = []
        for i in range(len(op_income)):
            oi = op_income[i] if i < len(op_income) else None
            rev = revenue[i] if i < len(revenue) else None
            if oi is not None and rev and rev != 0:
                ts_data["op_margin"].append(round(oi / rev * 100, 2))
            else:
                ts_data["op_margin"].append(None)

    # DuPont分解
    max_len = len(revenue) if revenue else 0
    total_assets_list = g_list('total_assets')
    total_equity_list = g_list('total_equity')
    net_margin_ts = []
    asset_turnover_ts = []
    fin_leverage_ts = []
    for i in range(max_len):
        ni = net_income[i] if i < len(net_income) else None
        rev = revenue[i] if i < len(revenue) else None
        ta = total_assets_list[i] if i < len(total_assets_list) else None
        eq = total_equity_list[i] if i < len(total_equity_list) else None
        if ni is not None and rev and rev != 0:
            net_margin_ts.append(round(ni / rev * 100, 2))
        else:
            net_margin_ts.append(None)
        if rev is not None and ta and ta != 0:
            asset_turnover_ts.append(round(rev / ta, 3))
        else:
            asset_turnover_ts.append(None)
        if ta is not None and eq and eq != 0:
            fin_leverage_ts.append(round(ta / eq, 3))
        else:
            fin_leverage_ts.append(None)

    ts_data["net_margin"] = net_margin_ts
    ts_data["asset_turnover"] = asset_turnover_ts
    ts_data["financial_leverage"] = fin_leverage_ts
    ts_data["interest_burden"] = []
    ts_data["nonop_burden"] = []
    ts_data["tax_burden"] = []

    return data, ts_data


# ---------- 既存のヘルパー関数 ----------

def _get_row_data(ws, row_label):
    """指定ラベルの行データを取得（新しい順）"""
    if ws is None:
        return []
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == row_label:
            vals = []
            for c in range(2, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    vals.append(v)
            return list(reversed(vals))  # 新しい順
    return []


def _safe_get(lst, idx, default=None):
    return lst[idx] if idx < len(lst) else default


def _find_sheet(wb, candidates):
    """候補名リストからシートを探す。見つからなければNone"""
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    # 部分一致でも探す
    lower_sheets = {s.lower(): s for s in wb.sheetnames}
    for name in candidates:
        for key, real_name in lower_sheets.items():
            if name.lower() in key:
                return wb[real_name]
    return None


def _try_labels(ws, labels):
    """複数のラベル候補から最初にヒットしたデータを返す"""
    for label in labels:
        data = _get_row_data(ws, label)
        if data:
            return data
    return []


# ---------- メインパース関数 ----------

def parse_excel(filepath):
    """Excelファイルをパースしてanalyzer用のdictを返す。
    .xls/.xlsx両対応。日本語縦型レイアウトも自動検出する。"""
    wb = _load_workbook(filepath)

    # 日本語縦型レイアウトの検出
    if _is_japanese_vertical_layout(wb):
        return _parse_japanese_vertical(wb)

    # --- 以下、従来の英語マルチシート形式 ---

    # シートの自動検出（複数の名前パターンに対応）
    inc = _find_sheet(wb, ['Income-Annual', 'Income Statement', 'Income', 'Export', 'income'])
    bs = _find_sheet(wb, ['Balance-Sheet-Annual', 'Balance Sheet', 'Balance', 'balance'])
    cf = _find_sheet(wb, ['Cash-Flow-Annual', 'Cash Flow', 'Cash Flow Statement', 'cashflow'])
    rat = _find_sheet(wb, ['Ratios-Annual', 'Ratios', 'Financial Ratios', 'ratios'])

    # もし全シートNoneなら、最初のシートをincとして使う
    if inc is None and bs is None and cf is None and rat is None:
        if wb.sheetnames:
            inc = wb[wb.sheetnames[0]]

    # 日付（新しい順）- 複数のラベルに対応
    dates_raw = _try_labels(inc, ['Date', 'Year Ending', 'Fiscal Year', 'Period'])

    # 収益データ
    revenue = _try_labels(inc, ['Revenue', 'Total Revenue', 'Net Revenue', 'Sales'])
    cogs_list = _try_labels(inc, ['Cost of Revenue', 'Cost of Goods Sold', 'COGS'])
    op_income = _try_labels(inc, ['Operating Income', 'Operating Profit', 'EBIT'])
    net_income = _try_labels(inc, ['Net Income', 'Net Profit', 'Net Earnings'])
    eps_list = _try_labels(inc, ['EPS (Basic)', 'EPS', 'Earnings Per Share'])
    op_margin_list = _try_labels(inc, ['Operating Margin'])
    ebitda_margin_list = _try_labels(inc, ['EBITDA Margin'])
    ebitda_list = _try_labels(inc, ['EBITDA'])
    sga_list = _try_labels(inc, ['Selling, General & Admin', 'SG&A', 'SGA'])
    da_list = _try_labels(inc, ['Depreciation & Amortization', 'D&A', 'Depreciation'])
    interest_exp_list = _try_labels(inc, ['Interest Expense / Income', 'Interest Expense', 'Net Interest Income'])
    other_exp_list = _try_labels(inc, ['Other Expense / Income', 'Other Income/Expense', 'Non-Operating Income'])
    pretax_income_list = _try_labels(inc, ['Pretax Income', 'Pre-Tax Income', 'Income Before Tax'])
    income_tax_list = _try_labels(inc, ['Income Tax', 'Tax Provision', 'Provision for Income Taxes'])
    eff_tax_rate_list = _try_labels(inc, ['Effective Tax Rate'])

    # キャッシュフロー
    fcf_list = _try_labels(cf, ['Free Cash Flow', 'FCF'])
    ocf_list = _try_labels(cf, ['Operating Cash Flow', 'Cash from Operations'])
    capex_list = _try_labels(cf, ['Capital Expenditures', 'Capex', 'CapEx'])
    investing_cf_list = _try_labels(cf, ['Investing Cash Flow', 'Cash from Investing'])
    financing_cf_list = _try_labels(cf, ['Financing Cash Flow', 'Cash from Financing'])

    # バランスシート
    total_assets_list = _try_labels(bs, ['Total Assets'])
    total_equity_list = _try_labels(bs, ['Shareholders Equity', "Shareholders' Equity", 'Total Equity', 'Stockholders Equity'])
    total_debt_list = _try_labels(bs, ['Total Debt', 'Long-Term Debt'])
    receivables_list = _try_labels(bs, ['Receivables', 'Accounts Receivable', 'Trade Receivables'])
    inventory_list = _try_labels(bs, ['Inventory', 'Inventories'])
    payables_list = _try_labels(bs, ['Accounts Payable', 'Trade Payables'])
    current_assets_list = _try_labels(bs, ['Total Current Assets', 'Current Assets'])
    current_liab_list = _try_labels(bs, ['Total Current Liabilities', 'Current Liabilities'])
    cash_list = _try_labels(bs, ['Cash & Cash Equivalents', 'Cash and Equivalents', 'Cash'])
    fixed_assets_list = _try_labels(bs, ['Property, Plant & Equipment', 'PP&E', 'Fixed Assets'])
    intangibles_list = _try_labels(bs, ['Goodwill and Intangibles', 'Intangible Assets', 'Goodwill'])
    net_debt_list = _try_labels(bs, ['Net Cash (Debt)', 'Net Debt'])
    long_term_assets_list = _try_labels(bs, ['Total Long-Term Assets', 'Non-Current Assets'])

    # 比率
    pe_list = _try_labels(rat, ['PE Ratio', 'P/E Ratio', 'Price/Earnings'])
    pb_list = _try_labels(rat, ['PB Ratio', 'P/B Ratio', 'Price/Book'])
    ev_list = _try_labels(rat, ['Enterprise Value', 'EV'])
    roe_list = _try_labels(rat, ['Return on Equity (ROE)', 'ROE', 'Return on Equity'])
    roa_list = _try_labels(rat, ['Return on Assets (ROA)', 'ROA', 'Return on Assets'])
    current_ratio_list = _try_labels(rat, ['Current Ratio'])
    quick_ratio_list = _try_labels(rat, ['Quick Ratio'])
    debt_fcf_list = _try_labels(rat, ['Debt/FCF'])
    debt_ebitda_list = _try_labels(rat, ['Debt/EBITDA'])
    nd_ebitda_list = _try_labels(rat, ['Net Debt/EBITDA'])
    roic_list = _try_labels(rat, ['Return on Invested Capital (ROIC)', 'ROIC'])
    debt_equity_list = _try_labels(rat, ['Debt/Equity', 'D/E Ratio'])
    dividend_yield_list = _try_labels(rat, ['Dividend Yield'])
    payout_ratio_list = _try_labels(rat, ['Payout Ratio'])

    # ROE/ROAがRatiosシートにないがIncome+BSから計算可能な場合
    if not roe_list and net_income and total_equity_list:
        roe_list = []
        max_len_roe = min(len(net_income), len(total_equity_list))
        for i in range(max_len_roe):
            ni = _safe_get(net_income, i)
            eq = _safe_get(total_equity_list, i)
            if ni is not None and eq and eq != 0:
                roe_list.append(ni / eq)
            else:
                roe_list.append(None)

    if not roa_list and net_income and total_assets_list:
        roa_list = []
        max_len_roa = min(len(net_income), len(total_assets_list))
        for i in range(max_len_roa):
            ni = _safe_get(net_income, i)
            ta = _safe_get(total_assets_list, i)
            if ni is not None and ta and ta != 0:
                roa_list.append(ni / ta)
            else:
                roa_list.append(None)

    # 値の取得ヘルパー
    def g(lst, idx):
        return _safe_get(lst, idx)

    # パーセント変換
    def to_pct(v):
        return v * 100 if v is not None else None

    # 自己資本比率の計算 D/Eベース: 1/(1+D/E)*100  (業種を問わず比較可能)
    equity_ratio = None
    equity_ratio_5y = None
    de0 = g(debt_equity_list, 0)
    de4 = g(debt_equity_list, 4)
    if de0 is not None:
        equity_ratio = (1 / (1 + de0)) * 100
    elif g(total_equity_list, 0) and g(total_assets_list, 0):
        equity_ratio = (g(total_equity_list, 0) / g(total_assets_list, 0)) * 100
    if de4 is not None:
        equity_ratio_5y = (1 / (1 + de4)) * 100
    elif g(total_equity_list, 4) and g(total_assets_list, 4):
        equity_ratio_5y = (g(total_equity_list, 4) / g(total_assets_list, 4)) * 100

    # 当座比率 (現在値 + 5年前)
    quick_r = to_pct(g(quick_ratio_list, 0))
    quick_r_5y = to_pct(g(quick_ratio_list, 4))

    # 流動比率 (現在値 + 5年前)
    current_r = to_pct(g(current_ratio_list, 0))
    current_r_5y = to_pct(g(current_ratio_list, 4))

    # 営業利益率 （%単位に変換）— 5年分
    op_margin_vals = [to_pct(g(op_margin_list, i)) for i in range(min(5, len(op_margin_list)))]
    # 営業利益率がRatiosにない場合、計算で補完
    if not op_margin_vals and op_income and revenue:
        op_margin_vals = []
        for i in range(min(5, len(op_income))):
            oi = g(op_income, i)
            rev = g(revenue, i)
            if oi is not None and rev and rev != 0:
                op_margin_vals.append(round(oi / rev * 100, 2))
            else:
                op_margin_vals.append(None)

    # EBITDAマージン (5年前)
    ebitda_margin_5y = to_pct(g(ebitda_margin_list, 4))

    # EBITDAマージン
    ebitda_margin_val = to_pct(g(ebitda_margin_list, 0))

    # ROE, ROA（%単位に変換）
    roe_now = to_pct(g(roe_list, 0))
    roe_3y = to_pct(g(roe_list, 2))
    roe_5y = to_pct(g(roe_list, 4))
    roa_now = to_pct(g(roa_list, 0))
    roa_3y = to_pct(g(roa_list, 2))
    roa_5y = to_pct(g(roa_list, 4))

    # ROE成長率
    roe_growth = roe_now - roe_5y if (roe_now is not None and roe_5y is not None) else None

    # NOPAT = Operating Income * (1 - Tax Rate)
    nopat = g(op_income, 0) * 0.75 if g(op_income, 0) else None
    nopat_5y = g(op_income, 4) * 0.75 if g(op_income, 4) else None

    # 投下資本 = Total Equity + Total Debt - Cash
    def calc_ic(eq, debt, cash):
        if eq is not None and debt is not None and cash is not None:
            return eq + debt - cash
        return None

    ic = calc_ic(g(total_equity_list, 0), g(total_debt_list, 0), g(cash_list, 0))
    ic_5y = calc_ic(g(total_equity_list, 4), g(total_debt_list, 4), g(cash_list, 4))

    # WACC簡易推定
    wacc_val = None
    if g(total_equity_list, 0) and g(total_debt_list, 0) and g(total_assets_list, 0):
        d_ratio = g(total_debt_list, 0) / (g(total_equity_list, 0) + g(total_debt_list, 0))
        e_ratio = 1 - d_ratio
        cost_of_equity = 8.0
        cost_of_debt = 3.0
        tax_rate = 0.25
        wacc_val = e_ratio * cost_of_equity + d_ratio * cost_of_debt * (1 - tax_rate)

    # SGA比率
    sga_ratio = None
    sga_ratio_5y = None
    if g(sga_list, 0) and g(revenue, 0):
        sga_ratio = (g(sga_list, 0) / g(revenue, 0)) * 100
    if g(sga_list, 4) and g(revenue, 4):
        sga_ratio_5y = (g(sga_list, 4) / g(revenue, 4)) * 100

    data = {
        "company": "",
        "ticker": "",
        "industry": "製造・サービス",

        "revenue": [g(revenue, i) for i in range(min(5, len(revenue)))],
        "fcf": [g(fcf_list, i) for i in range(min(5, len(fcf_list)))],
        "eps": [g(eps_list, i) for i in range(min(5, len(eps_list)))],

        "roe": [roe_now, roe_3y, roe_5y],
        "roe_growth_rate": roe_growth,
        "roa": [roa_now, roa_3y, roa_5y],

        "equity_ratio": equity_ratio,
        "equity_ratio_5y": equity_ratio_5y,
        "quick_ratio": quick_r,
        "quick_ratio_5y": quick_r_5y,
        "current_ratio": current_r,
        "current_ratio_5y": current_r_5y,

        "operating_cf": [g(ocf_list, i) for i in range(min(5, len(ocf_list)))],
        "investing_cf": [g(investing_cf_list, i) for i in range(min(5, len(investing_cf_list)))],
        "financing_cf": [g(financing_cf_list, i) for i in range(min(5, len(financing_cf_list)))],
        "op_margin": op_margin_vals,
        "ebitda_margin": ebitda_margin_val,
        "ebitda_margin_5y": ebitda_margin_5y,

        "debt_fcf": g(debt_fcf_list, 0),
        "debt_fcf_5y": g(debt_fcf_list, 4),
        "nd_ebitda": g(nd_ebitda_list, 0),
        "ev": g(ev_list, 0),
        "per": g(pe_list, 0),
        "per_5y": g(pe_list, 4),
        "pbr": g(pb_list, 0),
        "pbr_5y": g(pb_list, 4),

        "nopat": nopat,
        "nopat_5y": nopat_5y,
        "invested_capital": ic,
        "invested_capital_5y": ic_5y,
        "wacc": wacc_val,

        "accounts_receivable": g(receivables_list, 0),
        "accounts_receivable_5y": g(receivables_list, 4),
        "inventory": g(inventory_list, 0),
        "inventory_5y": g(inventory_list, 4),
        "accounts_payable": g(payables_list, 0),
        "accounts_payable_5y": g(payables_list, 4),
        "cogs": g(cogs_list, 0),
        "cogs_5y": g(cogs_list, 4),
        "sga_ratio": sga_ratio,
        "sga_ratio_5y": sga_ratio_5y,

        "total_assets": g(total_assets_list, 0),
        "total_assets_5y": g(total_assets_list, 4),
        "fixed_assets": g(fixed_assets_list, 0),
        "fixed_assets_5y": g(fixed_assets_list, 4),
        "tangible_fixed_assets": g(fixed_assets_list, 0),
        "tangible_fixed_assets_5y": g(fixed_assets_list, 4),
        "intangible_fixed_assets": g(intangibles_list, 0),
        "intangible_fixed_assets_5y": g(intangibles_list, 4),

        "net_income_val": g(net_income, 0),
        "net_income_val_5y": g(net_income, 4),
        "op_income_val": g(op_income, 0),
        "op_income_val_5y": g(op_income, 4),
        "interest_exp": g(interest_exp_list, 0),
        "interest_exp_5y": g(interest_exp_list, 4),
        "other_exp": g(other_exp_list, 0),
        "other_exp_5y": g(other_exp_list, 4),
        "pretax_income": g(pretax_income_list, 0),
        "pretax_income_5y": g(pretax_income_list, 4),
        "income_tax": g(income_tax_list, 0),
        "income_tax_5y": g(income_tax_list, 4),
        "total_equity": g(total_equity_list, 0),
        "total_equity_5y": g(total_equity_list, 4),

        "dividend_yield": to_pct(g(dividend_yield_list, 0)),
        "dividend_yield_5y": to_pct(g(dividend_yield_list, 4)),
        "payout_ratio": to_pct(g(payout_ratio_list, 0)),
        "payout_ratio_5y": to_pct(g(payout_ratio_list, 4)),

        "d1_mgmt_change": "○",
        "d2_ownership": "○",
        "d3_esg": "○",
    }

    # 時系列データ（チャート用）
    ts_data = {
        "dates": [str(d)[:4] if d else "" for d in dates_raw],
        "revenue": [g(revenue, i) for i in range(len(revenue))],
        "net_income": [g(net_income, i) for i in range(len(net_income))],
        "fcf": [g(fcf_list, i) for i in range(len(fcf_list))],
        "eps": [g(eps_list, i) for i in range(len(eps_list))],
        "ocf": [g(ocf_list, i) for i in range(len(ocf_list))],
        "investing_cf": [g(investing_cf_list, i) for i in range(len(investing_cf_list))],
        "financing_cf": [g(financing_cf_list, i) for i in range(len(financing_cf_list))],
        "ebitda": [g(ebitda_list, i) for i in range(len(ebitda_list))],
        "total_assets": [g(total_assets_list, i) for i in range(len(total_assets_list))],
        "total_equity": [g(total_equity_list, i) for i in range(len(total_equity_list))],
        "total_debt": [g(total_debt_list, i) for i in range(len(total_debt_list))],
        "roe": [to_pct(g(roe_list, i)) for i in range(len(roe_list))],
        "roa": [to_pct(g(roa_list, i)) for i in range(len(roa_list))],
        "op_margin": [to_pct(g(op_margin_list, i)) for i in range(len(op_margin_list))],
        "quick_ratio": [to_pct(g(quick_ratio_list, i)) for i in range(len(quick_ratio_list))],
        "current_ratio": [to_pct(g(current_ratio_list, i)) for i in range(len(current_ratio_list))],
        "equity_ratio": [((1 / (1 + g(debt_equity_list, i))) * 100) if (g(debt_equity_list, i) is not None) else ((g(total_equity_list, i) / g(total_assets_list, i) * 100) if (g(total_equity_list, i) and g(total_assets_list, i)) else None) for i in range(max(len(debt_equity_list), len(total_equity_list)))],
        "ebitda_margin": [to_pct(g(ebitda_margin_list, i)) for i in range(len(ebitda_margin_list))],
        "debt_fcf": [g(debt_fcf_list, i) for i in range(len(debt_fcf_list))],
        "roic": [to_pct(g(roic_list, i)) for i in range(len(roic_list))],
        "capex": [g(capex_list, i) for i in range(len(capex_list))],
        "sga": [g(sga_list, i) for i in range(len(sga_list))],
        "da": [g(da_list, i) for i in range(len(da_list))],
        "pe_ratio": [g(pe_list, i) for i in range(len(pe_list))],
        "pb_ratio": [g(pb_list, i) for i in range(len(pb_list))],
        "debt_ebitda": [g(debt_ebitda_list, i) for i in range(len(debt_ebitda_list))],
        "nd_ebitda": [g(nd_ebitda_list, i) for i in range(len(nd_ebitda_list))],
        "dividend_yield": [to_pct(g(dividend_yield_list, i)) for i in range(len(dividend_yield_list))],
        "payout_ratio": [to_pct(g(payout_ratio_list, i)) for i in range(len(payout_ratio_list))],
    }

    # 営業利益率の時系列が空で、計算可能な場合
    if not ts_data["op_margin"] and op_income and revenue:
        ts_data["op_margin"] = []
        for i in range(len(op_income)):
            oi = g(op_income, i)
            rev = g(revenue, i)
            if oi is not None and rev and rev != 0:
                ts_data["op_margin"].append(round(oi / rev * 100, 2))
            else:
                ts_data["op_margin"].append(None)

    # DuPont分解の時系列を計算: ROE = 純利益率 × 総資産回転率 × 財務レバレッジ
    max_len = len(revenue) if revenue else 0
    net_margin_ts = []
    asset_turnover_ts = []
    fin_leverage_ts = []
    for i in range(max_len):
        ni = g(net_income, i)
        rev = g(revenue, i)
        ta = g(total_assets_list, i)
        eq = g(total_equity_list, i)
        # 純利益率 (%) = Net Income / Revenue * 100
        if ni is not None and rev and rev != 0:
            net_margin_ts.append(round(ni / rev * 100, 2))
        else:
            net_margin_ts.append(None)
        # 総資産回転率 (x) = Revenue / Total Assets
        if rev is not None and ta and ta != 0:
            asset_turnover_ts.append(round(rev / ta, 3))
        else:
            asset_turnover_ts.append(None)
        # 財務レバレッジ (x) = Total Assets / Equity
        if ta is not None and eq and eq != 0:
            fin_leverage_ts.append(round(ta / eq, 3))
        else:
            fin_leverage_ts.append(None)

    ts_data["net_margin"] = net_margin_ts
    ts_data["asset_turnover"] = asset_turnover_ts
    ts_data["financial_leverage"] = fin_leverage_ts

    # 純利益率の分解時系列: 金利負担率・営業外損益率・税引後利益率
    interest_burden_ts = []
    tax_burden_ts = []
    nonop_burden_ts = []
    for i in range(max_len):
        oi = g(op_income, i)
        pt = g(pretax_income_list, i)
        ni = g(net_income, i)
        ie = g(interest_exp_list, i)
        oe = g(other_exp_list, i)

        if oi is not None and oi != 0 and ie is not None:
            interest_burden_ts.append(round((oi + ie) / oi * 100, 2))
        elif oi is not None and oi != 0 and pt is not None:
            interest_burden_ts.append(None)
        else:
            interest_burden_ts.append(None)

        if oi is not None and ie is not None and (oi + ie) != 0 and pt is not None:
            nonop_burden_ts.append(round(pt / (oi + ie) * 100, 2))
        else:
            nonop_burden_ts.append(None)

        if pt is not None and pt != 0 and ni is not None:
            tax_burden_ts.append(round(ni / pt * 100, 2))
        else:
            tax_burden_ts.append(None)

    ts_data["interest_burden"] = interest_burden_ts
    ts_data["nonop_burden"] = nonop_burden_ts
    ts_data["tax_burden"] = tax_burden_ts

    return data, ts_data


# ---------- カスタム分析: 可視化可能データのスキャン ----------

# メトリクス定義: key, Excelラベル, シート名, カテゴリ, 単位, 表示名(ja), 表示名(en)
METRIC_CATALOG = [
    {"key": "revenue",          "sheet": "inc", "label": "Revenue",                    "cat": "growth",      "unit": "百万", "ja": "売上高",              "en": "Revenue"},
    {"key": "op_income",        "sheet": "inc", "label": "Operating Income",            "cat": "profitability","unit": "百万", "ja": "営業利益",            "en": "Operating Income"},
    {"key": "ebitda",           "sheet": "inc", "label": "EBITDA",                      "cat": "profitability","unit": "百万", "ja": "EBITDA",              "en": "EBITDA"},
    {"key": "net_income",       "sheet": "inc", "label": "Net Income",                 "cat": "profitability","unit": "百万", "ja": "純利益",              "en": "Net Income"},
    {"key": "eps",              "sheet": "inc", "label": "EPS (Basic)",                "cat": "growth",      "unit": "円",   "ja": "EPS",                 "en": "EPS"},
    {"key": "op_margin",        "sheet": "inc", "label": "Operating Margin",           "cat": "profitability","unit": "%",    "ja": "営業利益率",          "en": "Operating Margin"},
    {"key": "ebitda_margin",    "sheet": "inc", "label": "EBITDA Margin",              "cat": "profitability","unit": "%",    "ja": "EBITDAマージン",      "en": "EBITDA Margin"},
    {"key": "sga",              "sheet": "inc", "label": "Selling, General & Admin",   "cat": "other",       "unit": "百万", "ja": "販管費",              "en": "SG&A"},
    {"key": "da",               "sheet": "inc", "label": "Depreciation & Amortization","cat": "other",       "unit": "百万", "ja": "減価償却費",          "en": "D&A"},
    {"key": "cogs",             "sheet": "inc", "label": "Cost of Revenue",            "cat": "other",       "unit": "百万", "ja": "売上原価",            "en": "COGS"},
    {"key": "fcf",              "sheet": "cf",  "label": "Free Cash Flow",             "cat": "health",      "unit": "百万", "ja": "FCF",                 "en": "FCF"},
    {"key": "ocf",              "sheet": "cf",  "label": "Operating Cash Flow",        "cat": "health",      "unit": "百万", "ja": "営業CF",              "en": "Operating CF"},
    {"key": "capex",            "sheet": "cf",  "label": "Capital Expenditures",       "cat": "other",       "unit": "百万", "ja": "設備投資",            "en": "CapEx"},
    {"key": "investing_cf",     "sheet": "cf",  "label": "Investing Cash Flow",        "cat": "health",      "unit": "百万", "ja": "投資CF",              "en": "Investing CF"},
    {"key": "financing_cf",     "sheet": "cf",  "label": "Financing Cash Flow",        "cat": "health",      "unit": "百万", "ja": "財務CF",              "en": "Financing CF"},
    {"key": "total_assets",     "sheet": "bs",  "label": "Total Assets",               "cat": "health",      "unit": "百万", "ja": "総資産",              "en": "Total Assets"},
    {"key": "total_equity",     "sheet": "bs",  "label": "Shareholders Equity",        "cat": "health",      "unit": "百万", "ja": "自己資本",            "en": "Equity"},
    {"key": "total_debt",       "sheet": "bs",  "label": "Total Debt",                 "cat": "health",      "unit": "百万", "ja": "有利子負債",          "en": "Total Debt"},
    {"key": "cash",             "sheet": "bs",  "label": "Cash & Cash Equivalents",    "cat": "health",      "unit": "百万", "ja": "現金",                "en": "Cash"},
    {"key": "receivables",      "sheet": "bs",  "label": "Receivables",                "cat": "other",       "unit": "百万", "ja": "売上債権",            "en": "Receivables"},
    {"key": "inventory",        "sheet": "bs",  "label": "Inventory",                  "cat": "other",       "unit": "百万", "ja": "棚卸資産",            "en": "Inventory"},
    {"key": "current_assets",   "sheet": "bs",  "label": "Total Current Assets",       "cat": "health",      "unit": "百万", "ja": "流動資産",            "en": "Current Assets"},
    {"key": "current_liab",     "sheet": "bs",  "label": "Total Current Liabilities",  "cat": "health",      "unit": "百万", "ja": "流動負債",            "en": "Current Liabilities"},
    {"key": "pe_ratio",         "sheet": "rat", "label": "PE Ratio",                   "cat": "valuation",   "unit": "x",    "ja": "PER",                 "en": "P/E Ratio"},
    {"key": "pb_ratio",         "sheet": "rat", "label": "PB Ratio",                   "cat": "valuation",   "unit": "x",    "ja": "PBR",                 "en": "P/B Ratio"},
    {"key": "roe",              "sheet": "rat", "label": "Return on Equity (ROE)",     "cat": "performance", "unit": "%",     "ja": "ROE",                 "en": "ROE"},
    {"key": "roa",              "sheet": "rat", "label": "Return on Assets (ROA)",     "cat": "performance", "unit": "%",     "ja": "ROA",                 "en": "ROA"},
    {"key": "roic",             "sheet": "rat", "label": "Return on Invested Capital (ROIC)","cat": "performance","unit": "%","ja": "ROIC",                "en": "ROIC"},
    {"key": "dividend_yield",   "sheet": "rat", "label": "Dividend Yield",                "cat": "performance", "unit": "%",     "ja": "配当利回り",          "en": "Dividend Yield"},
    {"key": "payout_ratio",     "sheet": "rat", "label": "Payout Ratio",                  "cat": "performance", "unit": "%",     "ja": "配当性向",            "en": "Payout Ratio"},
    {"key": "current_ratio",    "sheet": "rat", "label": "Current Ratio",              "cat": "health",      "unit": "%",     "ja": "流動比率",            "en": "Current Ratio"},
    {"key": "quick_ratio",      "sheet": "rat", "label": "Quick Ratio",                "cat": "health",      "unit": "%",     "ja": "当座比率",            "en": "Quick Ratio"},
    {"key": "debt_ebitda",      "sheet": "rat", "label": "Debt/EBITDA",                "cat": "health",      "unit": "x",     "ja": "Debt/EBITDA",         "en": "Debt/EBITDA"},
    {"key": "nd_ebitda",        "sheet": "rat", "label": "Net Debt/EBITDA",            "cat": "valuation",   "unit": "x",     "ja": "Net Debt/EBITDA",     "en": "Net Debt/EBITDA"},
    {"key": "debt_fcf",         "sheet": "rat", "label": "Debt/FCF",                   "cat": "health",      "unit": "x",     "ja": "Debt/FCF",            "en": "Debt/FCF"},
    {"key": "ev",               "sheet": "rat", "label": "Enterprise Value",           "cat": "valuation",   "unit": "百万",  "ja": "EV",                  "en": "Enterprise Value"},
    {"key": "fixed_assets",     "sheet": "bs",  "label": "Property, Plant & Equipment","cat": "other",       "unit": "百万",  "ja": "有形固定資産",        "en": "PP&E"},
    {"key": "intangibles",      "sheet": "bs",  "label": "Goodwill and Intangibles",   "cat": "other",       "unit": "百万",  "ja": "のれん・無形資産",    "en": "Goodwill & Intangibles"},
    {"key": "net_debt",         "sheet": "bs",  "label": "Net Cash (Debt)",            "cat": "health",      "unit": "百万",  "ja": "ネットキャッシュ",    "en": "Net Cash (Debt)"},
    {"key": "long_term_assets", "sheet": "bs",  "label": "Total Long-Term Assets",     "cat": "other",       "unit": "百万",  "ja": "固定資産",            "en": "Long-Term Assets"},
]


def scan_available_metrics(filepath):
    """Excelファイルをスキャンし、可視化可能なメトリクス一覧を返す"""
    wb = _load_workbook(filepath)

    # 日本語縦型レイアウトの場合は専用処理
    if _is_japanese_vertical_layout(wb):
        return _scan_japanese_metrics(wb)

    sheet_map = {
        'inc': _find_sheet(wb, ['Income-Annual', 'Income Statement', 'Income', 'Export', 'income']),
        'bs':  _find_sheet(wb, ['Balance-Sheet-Annual', 'Balance Sheet', 'Balance', 'balance']),
        'cf':  _find_sheet(wb, ['Cash-Flow-Annual', 'Cash Flow', 'Cash Flow Statement', 'cashflow']),
        'rat': _find_sheet(wb, ['Ratios-Annual', 'Ratios', 'Financial Ratios', 'ratios']),
    }

    available = []
    for m in METRIC_CATALOG:
        ws = sheet_map.get(m['sheet'])
        if ws is None:
            continue
        data = _get_row_data(ws, m['label'])
        if data:
            numeric = [v for v in data if isinstance(v, (int, float))]
            if numeric:
                available.append({
                    "key": m['key'],
                    "ja": m['ja'],
                    "en": m['en'],
                    "cat": m['cat'],
                    "unit": m['unit'],
                    "data_points": len(numeric),
                    "latest_value": numeric[0],
                })
    return available


def _scan_japanese_metrics(wb):
    """日本語縦型レイアウトからメトリクス一覧を返す"""
    data, _ = _parse_japanese_vertical(wb)
    # dataの各キーからMETRIC_CATALOGに対応するものを抽出
    jp_key_to_catalog = {
        'revenue': 'revenue', 'op_income': 'op_income', 'net_income': 'net_income',
        'eps': 'eps', 'ocf': 'ocf', 'investing_cf': 'investing_cf',
        'financing_cf': 'financing_cf', 'total_assets': 'total_assets',
        'total_equity': 'total_equity',
    }
    catalog_map = {m['key']: m for m in METRIC_CATALOG}
    available = []

    # data辞書からリスト形式のデータを検出
    check_keys = {
        'revenue': data.get('revenue', []),
        'op_income': [data.get('op_income_val')],
        'net_income': [data.get('net_income_val')],
        'eps': data.get('eps', []),
        'ocf': data.get('operating_cf', []),
        'investing_cf': data.get('investing_cf', []),
        'financing_cf': data.get('financing_cf', []),
        'total_assets': [data.get('total_assets')],
        'total_equity': [data.get('total_equity')],
        'roe': data.get('roe', []),
        'roa': data.get('roa', []),
        'payout_ratio': [data.get('payout_ratio')],
    }

    for key, vals in check_keys.items():
        m = catalog_map.get(key)
        if m is None:
            continue
        numeric = [v for v in vals if v is not None and isinstance(v, (int, float))]
        if numeric:
            available.append({
                "key": m['key'],
                "ja": m['ja'],
                "en": m['en'],
                "cat": m['cat'],
                "unit": m['unit'],
                "data_points": len(numeric),
                "latest_value": numeric[0],
            })
    return available


def extract_custom_timeseries(filepath, selected_keys):
    """選択されたメトリクスの時系列データを返す"""
    wb = _load_workbook(filepath)

    # 日本語縦型レイアウトの場合
    if _is_japanese_vertical_layout(wb):
        _, ts_data = _parse_japanese_vertical(wb)
        result = {"dates": ts_data.get("dates", [])}
        for key in selected_keys:
            if key in ts_data:
                result[key] = ts_data[key]
        return result

    sheet_map = {
        'inc': _find_sheet(wb, ['Income-Annual', 'Income Statement', 'Income', 'Export', 'income']),
        'bs':  _find_sheet(wb, ['Balance-Sheet-Annual', 'Balance Sheet', 'Balance', 'balance']),
        'cf':  _find_sheet(wb, ['Cash-Flow-Annual', 'Cash Flow', 'Cash Flow Statement', 'cashflow']),
        'rat': _find_sheet(wb, ['Ratios-Annual', 'Ratios', 'Financial Ratios', 'ratios']),
    }

    # 日付はincシートから取得
    inc_ws = sheet_map.get('inc')
    dates = _try_labels(inc_ws, ['Date', 'Year Ending', 'Fiscal Year', 'Period']) if inc_ws else []
    date_strs = [str(d)[:4] if d else "" for d in dates]

    result = {"dates": date_strs}
    catalog_map = {m['key']: m for m in METRIC_CATALOG}

    for key in selected_keys:
        m = catalog_map.get(key)
        if not m:
            continue
        ws = sheet_map.get(m['sheet'])
        if ws is None:
            continue
        data = _get_row_data(ws, m['label'])
        is_pct = m['unit'] == '%'
        if is_pct:
            data = [v * 100 if isinstance(v, (int, float)) else None for v in data]
        result[key] = data

    return result
